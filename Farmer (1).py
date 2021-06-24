s1 = 0
s2 = 0
o=0
result=0
land=10
import numpy as np
from openpyxl import load_workbook

income=np.zeros(6)
rf=np.zeros(5)
a=np.zeros(5)
indx=np.zeros(6)
collection=np.zeros(6,float)
available=np.zeros(6,float)
price=np.zeros(6,float)
profit=np.zeros(6,float)

wb = load_workbook('./FARMER.xlsx')
print(wb.sheetnames)
sheet = wb.get_sheet_by_name('Лист1')
for i in range (6):
    string=str(sheet.cell(row=17, column=i+2).value)
    string=string.replace(',','.')
    rf=float(string)
    collection[i]=float(string)
for i in range (6):
    string=str(sheet.cell(row=17, column=i+8).value)
    string=string.replace(',','.')
    rf=float(string)
    available[i]=rf
for i in range (6):
    string=str(sheet.cell(row=17, column=i+14).value)
    string=string.replace(',','.')
    rf=float(string)
    price[i]=rf



for i in range (6):
    profit[i]=price[i]*collection[i]

    
    
while land>0:
    m=np.argmax(profit)
    land=land-available[m]
    profit[m] = 0 
    if(land < 0):
        land = land + available[m]
        break 


for i in range (6):
    income[i]=price[i]*collection[i]

result = land*income[m]

for i in range (6):
    s1 = s1 + income[i]*available[i]
    
for i in range (6):
    s2 = s2 + profit[i]*available[i]

result = result + s1 - s2


print('maximum profit:', result)
